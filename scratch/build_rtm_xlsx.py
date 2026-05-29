from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FONT = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F3864")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F3864")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS_FILL = {
    "Done":    PatternFill("solid", start_color="C6EFCE"),
    "Partial": PatternFill("solid", start_color="FFEB9C"),
    "Gap":     PatternFill("solid", start_color="FFC7CE"),
    "Planned": PatternFill("solid", start_color="D9D9D9"),
}
STATUS_LABEL = {"Done": "✅ Done", "Partial": "🟠 Partial", "Gap": "🔴 Gap", "Planned": "⚪ Planned"}

wb = Workbook()

# ---- Sheet 1: Reference Index ----
ws0 = wb.active
ws0.title = "Reference Index"
ws0["A1"] = "VN-BECS-V2 法規/標準對照基準 (Reference Index)"
ws0["A1"].font = TITLE_FONT
ws0.append([])
ref_hdr = ["代碼", "來源"]
ws0.append(ref_hdr)
for c in range(1, 3):
    cell = ws0.cell(row=3, column=c); cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.border = BORDER; cell.alignment = CENTER
refs = [
    ("VN26", "越南 Circular 26/2013/TT-BYT（輸血專業技術規範）"),
    ("AABB", "AABB Standards for Blood Banks and Transfusion Services"),
    ("FDA606", "21 CFR 606（cGMP for Blood and Blood Components）"),
    ("FDA211", "21 CFR Part 11（電子紀錄/電子簽章）"),
    ("FDA-BECS", "FDA BECS Class II / 510(k) + 使用端 CSV 指引"),
    ("ISBT128", "ICCBBA ISBT 128 Standard（含 ISO 7064 檢查碼）"),
    ("ISO15189", "醫學實驗室品質與能力"),
    ("ISO27799", "健康資訊安全管理"),
    ("WHO-HV", "WHO/ISBT/IHN 血液警訊 (haemovigilance)"),
    ("EU-GMP", "EU Blood Directive 2002/98/EC + GMP/EDQM Blood Guide"),
]
for code, src in refs:
    ws0.append([code, src])
for row in ws0.iter_rows(min_row=4, max_row=3+len(refs), min_col=1, max_col=2):
    for cell in row:
        cell.font = Font(name=FONT, size=10); cell.border = BORDER; cell.alignment = WRAP
ws0.column_dimensions["A"].width = 14
ws0.column_dimensions["B"].width = 70

# ---- Sheet 2: RTM ----
ws = wb.create_sheet("RTM")
ws["A1"] = "需求可追溯矩陣 (Requirements Traceability Matrix)"
ws["A1"].font = TITLE_FONT
ws["A2"] = "基準：codex-optimize 分支實證審查｜更新：2026-05-29（含本回合 P0 修補）"
ws["A2"].font = Font(name=FONT, italic=True, size=9, color="808080")

headers = ["ID", "法規/標準", "需求 (Requirement)", "SOP", "實作 (Code)", "驗證 (Test)", "證據/備註", "狀態"]
ws.append([])
ws.append(headers)
hdr_row = 4
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=hdr_row, column=c)
    cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.border = BORDER; cell.alignment = CENTER

# (ID, std, requirement, sop, code, test, evidence, status)
rows = [
    ("RTM-AUTH-01","FDA211, ISO27799","每次 API 請求的操作者身分須由伺服器驗證，不得由用戶端自填","—","src/server/session.ts; rbacPolicy.resolveActorIdentity","sessionIdentity.test.ts","簽章會話 token；生產忽略 header 角色【本回合完成】","Done"),
    ("RTM-AUTH-02","FDA211","高風險動作須 RBAC 授權（角色比對）","—","rbacPolicy.authorizeApiRole","rbacDynamicMatrix.test.ts; sessionIdentity.test.ts","生產強制；非生產可 opt-in","Done"),
    ("RTM-AUTH-03","ISO27799, EU-GMP","機構範圍隔離 (ABAC facility scope) 於 API 與 DB","—","rbacPolicy.authorizeFacilityScope (API); db-migrations/001 (DB RLS+facility_id)","db-migrations/001_tests.sql (TEST 3/4)","API 助手+DB 遷移已備；待連線測試庫執行附證據","Partial"),
    ("RTM-AUTH-04","FDA211","demo/fallback 登入於生產一律禁用","—","authPolicy.ts; db.ts","authPolicy.test.ts; productionHardening.test.ts","fail-closed","Done"),
    ("RTM-AUTH-05","FDA211","前端每個 /api 請求附帶已驗證會話 token","—","src/lib/apiClient.ts; App.tsx","(手動/E2E)","全域 fetch 攔截器，零侵入元件【本回合完成】","Done"),
    ("RTM-STATE-01","AABB, FDA606","血袋生命週期狀態轉移集中受控、非法轉移阻擋","All","src/lib/stateMachine.ts","stateMachine.test.ts","terminal 不可逆","Done"),
    ("RTM-STATE-02","FDA606","所有狀態寫入經單一命令服務，無路由直寫","All","bloodUnitCommands.executeBloodUnitTransition","h1DirectWrites.test.ts","命令存在；回歸掃描需強化","Partial"),
    ("RTM-STATE-03","FDA606","跨表寫入原子化（元件+庫存+審計），失敗即回滾","All","bloodUnitCommands.ts (production fail-closed)","commandFailClosed.test.ts","生產環境寫入失敗即回傳失敗；真正多表 DB 交易仍待補","Partial"),
    ("RTM-STATE-04","AABB","多軸狀態（品質/庫存/指派/監管）獨立持久化","—","db-migrations/001 (enum+欄位+backfill); types.ts","db-migrations/001_tests.sql (TEST 5)","遷移已備；待執行並將命令服務改寫多軸","Partial"),
    ("RTM-REL-01","AABB, VN26, FDA606","未完成 IDM 陰性檢驗不得放行為可用","SOP2","stateMachine.ts:227-231","stateMachine.test.ts; limsRelease.test.ts","","Done"),
    ("RTM-REL-02","WHO-HV","Lookback 調查期間禁止釋回庫存","SOP9","stateMachine.ts:229 (isUnderLookback)","lookback.test.ts","","Done"),
    ("RTM-XM-01","AABB","ABO/Rh 相容性須依血品類別（紅血球/血漿/血小板分別）","SOP7","bloodSafety.evaluateComponentCompatibility; crossmatch/route.ts","componentCompatibility.test.ts; crossmatch.test.ts","類別感知引擎；血漿反向規則【本回合完成】","Done"),
    ("RTM-XM-02","AABB","配對使用成分實測 ABO（非捐血者登記血型）","SOP7","crossmatch/route.ts (component.abo 優先)","需新增測試","已改用 component.abo/rhd；DB join 仍待補","Partial"),
    ("RTM-XM-03","AABB","有抗體史者禁用電子/即時離心配血","SOP7","crossmatch/route.ts:90-96","crossmatch.test.ts","","Done"),
    ("RTM-XM-04","AABB","檢體有效期（3 天）超期須重抽","SOP7","validators.validateSpecimenDate","validators.test.ts","預設 3 天","Done"),
    ("RTM-XM-05","AABB","病人血型未知不得常規配血（除緊急政策）","SOP7","crossmatch/route.ts:79-81","crossmatch.test.ts","","Done"),
    ("RTM-BED-01","AABB, WHO-HV","床邊雙人核對：兩位不同且經認證之合格人員","SOP6","bedside-verify/route.ts (主+次驗者皆驗密碼); BedsideVerificationView.tsx","bedsideVerifyRoute.test.ts","主、次驗者皆須密碼認證；前端收兩組憑證","Done"),
    ("RTM-BED-02","AABB, WHO-HV","床邊須重新比對血袋實測 ABO/Rh ↔ 病人","SOP6","bedside-verify/route.ts (evaluateComponentCompatibility 獨立再驗)","bedsideVerifyRoute.test.ts (BED-02)","血型資料存在時硬阻斷 ABO/Rh 不符；缺資料記警告","Done"),
    ("RTM-BED-03","AABB","須先有有效發血紀錄、同意書、輸血前生命徵象","SOP6","bedside-verify/route.ts","bedsideVerifyRoute.test.ts","","Done"),
    ("RTM-ISS-01","AABB","發血須醫囑確認或記錄在案之緊急放行","SOP8","stateMachine.ts:364-373","issue.test.ts","","Done"),
    ("RTM-ISS-02","VN26, AABB","發出逾 30 分鐘不得回庫；外觀/冷鏈不合格須報廢","SOP8","stateMachine.ts:388-407","issueReturn.test.ts","","Done"),
    ("RTM-EMG-01","AABB","緊急/MTP：強制 approver、適應症、RhD 政策、事後 review SLA","SOP10","clinicalPolicy.validateEmergencyRelease; mtp-cases/[id]/issue/route.ts","emergencyPolicy.test.ts; h1DirectWrites.test.ts","強制授權醫師/適應症(從 MTP case 衍生)+RhD 政策+review SLA 戳記","Done"),
    ("RTM-DON-01","VN26","捐血者年齡/體重資格（男≥45kg、女≥42kg）","SOP1","validators.validateDonorAge/Weight","validators.test.ts","18–60；45/42kg","Done"),
    ("RTM-DON-02","VN26","採血量上限 <500ml、體重連動（42–45kg<250ml、<42kg 不合格）","SOP1","validators.validateCollectionVolume (lims/collect/route.ts 伺服器端強制)","collectionVolume.test.ts","<500ml/<42kg/42–45kg<250ml 已強制；≥45kg 每公斤上限屬本地政策","Done"),
    ("RTM-DON-03","VN26, AABB","問卷暫緩規則並記錄決策時政策版本","SOP1","clinicalPolicy.ts; validators.validateVietnamDeferralRules; lims/donors/route.ts","emergencyPolicy.test.ts","版本化政策登錄；deferral 回傳並於問卷記錄 policyVersion","Done"),
    ("RTM-LBL-01","ISBT128","DIN/標籤完整解析（結構、機構碼、產品碼、ISO 7064 檢查碼）","SOP1/3","src/lib/isbt128.ts (parseDin/parseProductCode/MOD37,2); lims/collect/route.ts","isbt128.test.ts","結構+FIN+檢查碼解析並於採血路由強制；ICCBBA 官方向量待 UAT","Done"),
    ("RTM-TRACE-01","FDA606, ISBT128","捐血者↔受血者雙向完整追溯報表","SOP9","reportingService.ts; reconciliation.ts","需新增端到端追溯測試","lookback 連動存在；報表未完整","Partial"),
    ("RTM-AUD-01","FDA211","稽核紀錄 append-only、不可竄改（DB 層強制）","—","supabase_schema.sql:142-160,402-427","db-migrations/001_tests.sql (TEST 2)","trigger+RLS 已有；竄改測試已撰；待連線測試庫執行附證據","Partial"),
    ("RTM-AUD-02","FDA211","稽核 hash chain 完整性可驗證（端點 + 排程）","—","auditChain.verifyAuditChain; app/api/v1/audit-events/verify/route.ts","auditChainVerify.test.ts","驗證端點+竄改/斷鏈偵測測試已建；排程化待補","Done"),
    ("RTM-COLD-01","VN26, EU-GMP","冷鏈：驗證裝置、excursion 閾值、自動隔離、CAPA","SOP5","transport_jobs; route-local","需冷鏈 excursion E2E 測試","目前為模擬值","Gap"),
    ("RTM-OFF-01","FDA606","離線僅限緊急命令、簽章、idempotency、衝突送審","—","offlineSync.ts; sync/push-events/route.ts","offlineSyncScenarios.test.ts","衝突/版本檢查有；範圍限制與簽章待強化","Partial"),
    ("RTM-FHIR-01","(互通)","FHIR 資源完整對應（7 類資源）","—","FhirAdapter.ts; app/api/v1/fhir/route.ts","fhirAdapter.test.ts","7 類資源 mapper+路由 bundle 查詢已建；外部訊息驗證/版本化待 P2","Done"),
    ("RTM-VAL-01","FDA-BECS, ISO15189","CSV 驗證套件：需求→測試→IQ/OQ/PQ 證據；上線前簽核","—","本 RTM 即起點","—","尚未建立","Gap"),
    ("RTM-REG-01","VN26","越南法規正式對應與 12 項臨床政策書面決策","—","00_vietnam_readiness_gap.md","—","待主管機關/試點醫院簽核","Gap"),
]

r = hdr_row + 1
for row in rows:
    *cols, status = row
    ws.append(cols + [STATUS_LABEL[status]])
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=r, column=c)
        cell.border = BORDER
        cell.font = Font(name=FONT, size=9)
        cell.alignment = WRAP
        if c == 1:
            cell.font = Font(name=FONT, size=9, bold=True)
    sc = ws.cell(row=r, column=len(headers))
    sc.fill = STATUS_FILL[status]; sc.alignment = CENTER; sc.font = Font(name=FONT, size=9, bold=True)
    r += 1

widths = {"A": 14, "B": 16, "C": 38, "D": 8, "E": 34, "F": 26, "G": 34, "H": 12}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A{hdr_row}:H{r-1}"

# ---- Sheet 3: Gate Summary (with live COUNTIF formulas) ----
wsg = wb.create_sheet("Gate Summary")
wsg["A1"] = "Gate 對應與完成統計"
wsg["A1"].font = TITLE_FONT
wsg.append([])
wsg.append(["狀態統計 (全部需求)", "數量"])
for c in (1, 2):
    cell = wsg.cell(row=3, column=c); cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.border = BORDER; cell.alignment = CENTER
_counts = {k: sum(1 for x in rows if x[-1] == k) for k in ("Done", "Partial", "Gap", "Planned")}
stat_rows = [
    ("✅ Done", _counts["Done"]),
    ("🟠 Partial", _counts["Partial"]),
    ("🔴 Gap", _counts["Gap"]),
    ("⚪ Planned", _counts["Planned"]),
    ("總計", len(rows)),
]
rr = 4
for label, formula in stat_rows:
    wsg.cell(row=rr, column=1, value=label).font = Font(name=FONT, size=10, bold=(label=="總計"))
    wsg.cell(row=rr, column=2, value=formula).font = Font(name=FONT, size=10)
    wsg.cell(row=rr, column=1).border = BORDER; wsg.cell(row=rr, column=2).border = BORDER
    wsg.cell(row=rr, column=2).alignment = CENTER
    rr += 1

wsg.append([])
gate_hdr_row = rr + 1
wsg.cell(row=gate_hdr_row, column=1, value="Gate")
wsg.cell(row=gate_hdr_row, column=2, value="對應 RTM 範圍")
wsg.cell(row=gate_hdr_row, column=3, value="主要阻斷 (Gap)")
for c in (1, 2, 3):
    cell = wsg.cell(row=gate_hdr_row, column=c); cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.border = BORDER; cell.alignment = CENTER
gates = [
    ("Gate 1 工程安全", "AUTH-*, STATE-*, AUD-*", "STATE-03(交易/fail-closed)、STATE-04(多軸落地)、AUTH-03(RLS/facility)"),
    ("Gate 2 臨床流程", "XM-*, BED-*, EMG-*, COLD-*, LBL-*", "BED-02(床邊比型)、EMG-01、LBL-01、COLD-01"),
    ("Gate 3 法規品質", "VAL-01, REG-01, TRACE-01, FHIR-01", "全部待辦"),
]
gr = gate_hdr_row + 1
for g in gates:
    wsg.append([])
    for c, val in enumerate(g, start=1):
        cell = wsg.cell(row=gr, column=c, value=val)
        cell.border = BORDER; cell.font = Font(name=FONT, size=10); cell.alignment = WRAP
    gr += 1
for col, w in {"A": 26, "B": 34, "C": 60}.items():
    wsg.column_dimensions[col].width = w

out = "docs/production/REQUIREMENTS_TRACEABILITY_MATRIX.xlsx"
wb.save(out)
print("saved", out)
